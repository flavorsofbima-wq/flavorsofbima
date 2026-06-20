"""
Build-time catalog generator.
Reads admin/FlavorsOfBIMA-Admin.xlsx and writes src/data/catalog.generated.json
which the site imports. Run automatically before dev/build via package.json.

WHY THIS EXISTS: a deployed site can't read an Excel file on your computer.
So we read it HERE (during build) and bake the result into the site.
Edit the Excel, run `npm run dev` (or redeploy) and changes appear.
"""
import json, sys, os
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "admin", "FlavorsOfBIMA-Admin.xlsx")
OUT = os.path.join(ROOT, "src", "data", "catalog.generated.json")

# Static metadata that doesn't live in the sheet (emoji, gradient, descriptions, ingredients)
# keyed by slug — merged with sheet data. This keeps rich content in code while
# prices/active/discounts come from the easily-edited sheet.
META = {
  "avakaya": {"emoji":"🥭","gradient":"linear-gradient(145deg,#1a0804,#5a1808,#9a2c10)","shortDesc":"Classic Andhra raw mango pickle with rich red masala.","description":"Our signature Avakaya — chunky raw mango pieces, fenugreek, mustard and rich red chilli masala, slow-set in wood pressed sesame oil. Bold, tangy and unmistakably homemade.","ingredients":["Raw mango","Red chilli powder","Mustard","Fenugreek","Wood pressed sesame oil","Salt","Garlic (optional)"],"badge":"🔥 #1 Seller"},
  "tomato": {"emoji":"🍅","gradient":"linear-gradient(145deg,#1a0808,#5a1818,#9a2828)","shortDesc":"Tangy-spicy homemade tomato pickle.","description":"Sun-ripened tomatoes cooked down with mustard, red chilli and Andhra spices into a thick, tangy pickle that pairs perfectly with rice, idly and dosa.","ingredients":["Tomato","Red chilli powder","Mustard","Tamarind","Wood pressed sesame oil","Salt","Garlic (optional)"],"badge":""},
  "gongura": {"emoji":"🌿","gradient":"linear-gradient(145deg,#0a1a0a,#1a5a1a,#2a8a2a)","shortDesc":"Quintessential Andhra sorrel leaf pickle.","description":"Tender gongura (sorrel) leaves slow-cooked with red chilli and spices into the tangy, bold pickle Andhra is famous for. A must with hot rice and ghee.","ingredients":["Gongura leaves","Red chilli powder","Mustard","Wood pressed sesame oil","Salt","Garlic (optional)"],"badge":"⭐ Most Ordered"},
  "red-chilli": {"emoji":"🌶️","gradient":"linear-gradient(145deg,#2a0808,#7a1010,#aa2020)","shortDesc":"Fiery whole red chilli pickle.","description":"For those who love the heat — plump red chillies stuffed with a mustard-forward masala and matured in wood pressed sesame oil. Intense, bold and traditional.","ingredients":["Red chilli","Mustard","Fenugreek","Wood pressed sesame oil","Salt"],"badge":""},
  "gongura-red-chilli": {"emoji":"🌿🌶️","gradient":"linear-gradient(145deg,#0d1a0d,#2a5a2a,#4a8a4a)","shortDesc":"Tangy gongura meets fiery red chilli.","description":"Best of both worlds — the tang of gongura blended with the heat of red chilli in one perfectly balanced, bold jar.","ingredients":["Gongura leaves","Red chilli","Mustard","Wood pressed sesame oil","Salt","Garlic (optional)"],"badge":"🌟 Mix"},
  "lemon": {"emoji":"🍋","gradient":"linear-gradient(145deg,#1a1a08,#585810,#8a8820)","shortDesc":"Bright, zesty lemon pickle.","description":"Whole lemon pieces matured with warming spices into a bright, tangy, slightly sweet-sour pickle that cuts through any rich meal.","ingredients":["Lemon","Red chilli powder","Mustard","Fenugreek","Wood pressed sesame oil","Salt"],"badge":""},
  "gooseberry": {"emoji":"🫐","gradient":"linear-gradient(145deg,#0a1a08,#285818,#408830)","shortDesc":"Healthy, tangy amla pickle.","description":"Nutrient-rich gooseberry (amla) pieces in a traditional Andhra spice blend — tangy, wholesome and full of goodness.","ingredients":["Gooseberry (amla)","Red chilli powder","Mustard","Wood pressed sesame oil","Salt"],"badge":""},
  "chicken-boneless": {"emoji":"🍗","gradient":"linear-gradient(145deg,#1a0808,#4a1010,#8a1a1a)","shortDesc":"Tender boneless chicken in Andhra spices.","description":"Small tender boneless chicken pieces cooked in an authentic Andhra spice blend — clearly visible meat, less masala, bold homemade taste. Set in wood pressed sesame oil.","ingredients":["Boneless chicken","Red chilli powder","Ginger-garlic","Andhra spices","Wood pressed sesame oil","Salt"],"badge":"🌟 Premium"},
  "gongura-chicken-boneless": {"emoji":"🍗🌿","gradient":"linear-gradient(145deg,#0d1a0d,#3a4a10,#6a7a1a)","shortDesc":"Boneless chicken in tangy gongura.","description":"Tender boneless chicken marinated with tangy gongura (sorrel) and Andhra spices — the ultimate non-veg pickle combination.","ingredients":["Boneless chicken","Gongura leaves","Red chilli powder","Ginger-garlic","Wood pressed sesame oil","Salt"],"badge":"Must Try"},
  "chicken-bone": {"emoji":"🍖","gradient":"linear-gradient(145deg,#1a0606,#4a0e0e,#8a1616)","shortDesc":"Bone-in chicken pickle, rich & traditional.","description":"Traditional country-style chicken pickle with bone-in cuts for deeper flavour, slow-cooked in an authentic Andhra spice blend and wood pressed sesame oil.","ingredients":["Chicken (with bone)","Red chilli powder","Ginger-garlic","Andhra spices","Wood pressed sesame oil","Salt"],"badge":"🍖 Traditional"},
  "gongura-chicken-bone": {"emoji":"🍖🌿","gradient":"linear-gradient(145deg,#0d1a0a,#3a4a0e,#6a7a16)","shortDesc":"Bone-in chicken in tangy gongura.","description":"Country-style bone-in chicken pickle blended with tangy gongura (sorrel) and Andhra spices — bold, rustic and full of flavour.","ingredients":["Chicken (with bone)","Gongura leaves","Red chilli powder","Ginger-garlic","Wood pressed sesame oil","Salt"],"badge":"Must Try"},
  "garam-masala": {"emoji":"🌶️","gradient":"linear-gradient(145deg,#1a0a04,#5a2a0a,#9a4810)","shortDesc":"Freshly ground whole spice blend.","description":"A warming, aromatic blend of whole spices roasted and ground fresh — the soul of Indian cooking.","ingredients":["Cardamom","Cinnamon","Clove","Cumin","Coriander","Black pepper"],"badge":""},
  "dhaniya-powder": {"emoji":"🌿","gradient":"linear-gradient(145deg,#1a1208,#5a4210,#9a7020)","shortDesc":"Roasted & ground coriander powder.","description":"Freshly roasted coriander seeds ground to a warm, citrusy powder that forms the base of countless curries.","ingredients":["Coriander seeds"],"badge":""},
  "jeera-powder": {"emoji":"🫘","gradient":"linear-gradient(145deg,#18140a,#504018,#887828)","shortDesc":"Hand-roasted cumin powder.","description":"Nutty, earthy hand-roasted cumin ground fresh — essential for tempering, raita and spice blends.","ingredients":["Cumin seeds"],"badge":""},
  "bisibele-bath-powder": {"emoji":"🍛","gradient":"linear-gradient(145deg,#1a0e08,#5a3210,#9a5820)","shortDesc":"Signature blend for bisibelebath.","description":"The complex, rich spice blend that gives Karnataka's famous bisibelebath its unmistakable flavour.","ingredients":["Chana dal","Coriander","Red chilli","Cinnamon","Clove","Dry coconut"],"badge":""},
  "vangi-bath-powder": {"emoji":"🍆","gradient":"linear-gradient(145deg,#0a1408,#2a4418,#4a7428)","shortDesc":"Aromatic blend for brinjal rice.","description":"A bold, nutty spice blend crafted for vangi bath (brinjal rice) — equally delicious with any vegetable rice.","ingredients":["Coriander","Chana dal","Urad dal","Red chilli","Dry coconut","Cinnamon"],"badge":""},
  "sambar-powder": {"emoji":"🥣","gradient":"linear-gradient(145deg,#1a0e06,#5a3410,#9a6020)","shortDesc":"Homemade sambar masala.","description":"A balanced, fragrant homemade sambar powder with the perfect depth for everyday sambar and stews.","ingredients":["Coriander","Toor dal","Chana dal","Red chilli","Fenugreek","Curry leaves"],"badge":""},
  "rasam-powder": {"emoji":"🍵","gradient":"linear-gradient(145deg,#16100a,#4a3410,#7a5a20)","shortDesc":"Light & aromatic rasam blend.","description":"A comforting, aromatic rasam powder that turns simple tamarind broth into soul-warming rasam.","ingredients":["Coriander","Cumin","Black pepper","Red chilli","Toor dal"],"badge":""},
  "curry-leaf-powder": {"emoji":"🌿","gradient":"linear-gradient(145deg,#0a1a0a,#1a4a1a,#2a7a2a)","shortDesc":"Fresh, greenish, aromatic podi.","description":"Fresh curry leaves roasted and ground with lentils and spices into a nutritious, fragrant podi. Greenish, premium and packed with flavour.","ingredients":["Curry leaves","Urad dal","Chana dal","Red chilli","Sesame","Salt"],"badge":"💚 Premium"},
  "gun-powder": {"emoji":"🌶️","gradient":"linear-gradient(145deg,#1a1008,#5a3010,#9a5020)","shortDesc":"Fiery, punchy idly-dosa podi.","description":"The classic fiery lentil podi — mix with ghee or oil and it transforms idly, dosa and plain rice into something unforgettable.","ingredients":["Urad dal","Chana dal","Red chilli","Sesame","Garlic","Salt"],"badge":"🔥 Spicy Hit"},
  "palli-karam-podi": {"emoji":"🥜","gradient":"linear-gradient(145deg,#1a0e08,#4a2a10,#8a4820)","shortDesc":"Roasted peanut karam podi.","description":"Roasted peanuts ground with red chilli and garlic into a fiery, nutty podi — a Telugu kitchen staple loved with rice and ghee.","ingredients":["Peanuts","Red chilli","Garlic","Cumin","Salt"],"badge":""},
  "vellulli-karam-podi": {"emoji":"🧄","gradient":"linear-gradient(145deg,#14100a,#3a3010,#6a5820)","shortDesc":"Bold garlic karam podi.","description":"An intensely flavoured garlic podi — pungent, rich and bold. A little goes a long way over rice, idly or dosa.","ingredients":["Garlic","Red chilli","Urad dal","Sesame","Salt"],"badge":""},
  "flax-seeds-podi": {"emoji":"🌾","gradient":"linear-gradient(145deg,#100e0a,#302e18,#5a5828)","shortDesc":"Nutritious roasted flax podi.","description":"Roasted flax seeds ground with mild spices into an omega-rich, healthy podi with a subtle nutty warmth.","ingredients":["Flax seeds","Red chilli","Garlic","Cumin","Salt"],"badge":""},
}

FALLBACK_CATEGORIES = [
  {"id":"veg-pickles","name":"Veg Pickles","icon":"🥭","type":"veg","base":"250g"},
  {"id":"non-veg-pickles","name":"Non-Veg Pickles","icon":"🍗","type":"non-veg","base":"250g"},
  {"id":"spices","name":"Spices","icon":"🌶️","type":"veg","base":"100g"},
  {"id":"podis","name":"Podis","icon":"🌿","type":"veg","base":"100g"},
]

def yn(v):
    return str(v).strip().lower() in ("yes","y","true","1")

def num(v):
    try:
        if v is None or str(v).strip()=="":
            return None
        return float(v)
    except: return None

def read_categories(wb):
    if "Categories" not in wb.sheetnames:
        return FALLBACK_CATEGORIES
    ws = wb["Categories"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hr = -1
    for i, r in enumerate(rows[:12]):
        low = [str(c).strip().lower() if c is not None else "" for c in r]
        if "id" in low and "name" in low:
            hr = i; break
    if hr == -1:
        return FALLBACK_CATEGORIES
    ch = {str(c).strip().lower(): i for i, c in enumerate(rows[hr]) if c}
    out = []
    for r in rows[hr+1:]:
        if not r or not r[ch["id"]]:
            continue
        cid = str(r[ch["id"]]).strip()
        if not cid:
            continue
        if "active" in ch and not yn(r[ch["active"]]):
            continue
        out.append({
            "id": cid,
            "name": str(r[ch["name"]]).strip(),
            "icon": str(r[ch["icon"]]).strip() if ch.get("icon") is not None and r[ch["icon"]] else "",
            "type": str(r[ch["type"]]).strip() if ch.get("type") is not None and r[ch["type"]] else "veg",
            "base": "100g" if str(r[ch["base_weight"]]).strip()=="100g" else "250g",
            "image": str(r[ch["image_filename"]]).strip() if ch.get("image_filename") is not None and r[ch["image_filename"]] else "",
        })
    return out or FALLBACK_CATEGORIES

def main():
    if not os.path.exists(XLSX):
        print(f"ERROR: {XLSX} not found", file=sys.stderr); sys.exit(1)
    wb = load_workbook(XLSX, data_only=True)
    CATEGORIES = read_categories(wb)

    # ---- Products ----
    ws = wb["Products"]
    # header row is row 4
    headers = [c.value for c in ws[4]]
    idx = {h: i for i, h in enumerate(headers) if h}
    products = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or not row[idx["slug"]]:
            continue
        slug = str(row[idx["slug"]]).strip()
        if not yn(row[idx["active"]]):
            continue  # item 12: only active products
        cat = str(row[idx["category"]]).strip()
        base = next((c["base"] for c in CATEGORIES if c["id"]==cat), "250g")
        # build weight->price map from the sheet
        prices = {}
        if base == "250g":
            for label, key in (("250g","price_250g"),("500g","price_500g"),("1kg","price_1kg")):
                p = num(row[idx[key]])
                if p is not None: prices[label] = round(p)
        else:
            for label, key in (("100g","price_100g"),("200g","price_200g"),("500g","price_500g")):
                if key in idx:
                    p = num(row[idx[key]])
                    if p is not None: prices[label] = round(p)
        meta = META.get(slug, {"emoji":"🫙","gradient":"linear-gradient(145deg,#0c2354,#071a3d)","shortDesc":"","description":"","ingredients":[],"badge":""})
        disc_type = (row[idx["discount_type"]] or "")
        disc_type = str(disc_type).strip().lower() if disc_type else ""
        disc_val = num(row[idx["discount_value"]]) or 0
        is_pickle = cat in ("veg-pickles", "non-veg-pickles")
        refined_diff = int(num(row[idx["refined_oil_diff"]]) or 0) if (is_pickle and "refined_oil_diff" in idx) else 0
        glass_diff = int(num(row[idx["glass_bottle_diff"]]) or 0) if (is_pickle and "glass_bottle_diff" in idx) else 0
        products.append({
            "slug": slug,
            "name": str(row[idx["name"]]).strip(),
            "category": cat,
            "prices": prices,
            "hasGarlic": yn(row[idx["has_garlic"]]),
            "garlicSurcharge": int(num(row[idx["garlic_surcharge"]]) or 0),
            "hasOptions": is_pickle,
            "refinedOilDiff": refined_diff,
            "glassBottleDiff": glass_diff,
            "discountType": disc_type if disc_type in ("amount","percentage") else "",
            "discountValue": disc_val,
            "bestseller": yn(row[idx["bestseller"]]),
            "seasonal": yn(row[idx["seasonal"]]),
            **meta,
        })

    # ---- Settings ----
    ws2 = wb["Settings"]
    settings = {}
    for row in ws2.iter_rows(min_row=4, values_only=True):
        if row and row[0]:
            settings[str(row[0]).strip()] = (str(row[1]).strip() if row[1] is not None else "")

    # ---- WhyUs ----
    ws3 = wb["WhyUs"]
    whyus = []
    for row in ws3.iter_rows(min_row=4, values_only=True):
        if row and row[1]:
            whyus.append({"order":num(row[0]) or 0,"title":str(row[1]).strip(),
                          "desc":str(row[2]).strip() if row[2] else "","image":str(row[3]).strip() if row[3] else ""})
    whyus.sort(key=lambda x: x["order"])

    # ---- Testimonials ----
    ws4 = wb["Testimonials"]
    testimonials = []
    for row in ws4.iter_rows(min_row=4, values_only=True):
        if row and row[1]:
            testimonials.append({"order":num(row[0]) or 0,"name":str(row[1]).strip(),
                "location":str(row[2]).strip() if row[2] else "","rating":int(num(row[3]) or 5),
                "review":str(row[4]).strip() if row[4] else "","image":str(row[5]).strip() if row[5] else ""})
    testimonials.sort(key=lambda x: x["order"])

    # ---- TrustBar ----
    trustbar = []
    if "TrustBar" in wb.sheetnames:
        for row in wb["TrustBar"].iter_rows(min_row=5, values_only=True):
            if row and row[1]:
                trustbar.append({"order":num(row[0]) or 0,"label":str(row[1]).strip(),
                    "emoji":str(row[2]).strip() if row[2] else "","image":str(row[3]).strip() if row[3] else ""})
        trustbar.sort(key=lambda x: x["order"])

    # ---- FAQ ----
    faq = []
    if "FAQ" in wb.sheetnames:
        for row in wb["FAQ"].iter_rows(min_row=4, values_only=True):
            if row and row[1]:
                faq.append({"order":num(row[0]) or 0,"q":str(row[1]).strip(),
                    "a":str(row[2]).strip() if row[2] else ""})
        faq.sort(key=lambda x: x["order"])

    import datetime
    out = {
        "generatedAt": datetime.datetime.now().isoformat(),
        "categories": CATEGORIES,
        "products": products,
        "settings": settings,
        "whyus": whyus,
        "testimonials": testimonials,
        "trustbar": trustbar,
        "faq": faq,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"✅ Generated {OUT}")
    print(f"   {len(products)} products, {len(CATEGORIES)} categories, {len(faq)} FAQs, {len(trustbar)} trust items")
    print(f"   fssai1={settings.get('fssai_license_1')}, fssai2={settings.get('fssai_license_2')}")

if __name__ == "__main__":
    main()
